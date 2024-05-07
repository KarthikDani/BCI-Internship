import customtkinter as ctk
import random
from openpyxl import load_workbook, Workbook

class UserForm(ctk.CTk):
    def __init__(self):
        super().__init__()

        # Configure the window
        self.title("User Details Form")
        self.geometry("500x600")

        # Customtkinter uses the `set_appearance_mode` function to set a theme
        ctk.set_appearance_mode("dark")  # Choose either "dark" or "light" mode

        # Create form widgets
        self.create_form()

    def create_form(self):
        # Create labels and entry fields
        self.name_label = ctk.CTkLabel(self, text="Name:")
        self.name_label.grid(row=0, column=0, sticky='w', padx=10, pady=10)
        self.name_entry = ctk.CTkEntry(self)
        self.name_entry.grid(row=0, column=1, padx=10, pady=10)

        self.age_label = ctk.CTkLabel(self, text="Age:")
        self.age_label.grid(row=1, column=0, sticky='w', padx=10, pady=10)
        self.age_entry = ctk.CTkEntry(self)
        self.age_entry.grid(row=1, column=1, padx=10, pady=10)

        self.gender_label = ctk.CTkLabel(self, text="Gender:")
        self.gender_label.grid(row=2, column=0, sticky='w', padx=10, pady=10)
        self.gender_entry = ctk.CTkEntry(self)
        self.gender_entry.grid(row=2, column=1, padx=10, pady=10)

        self.phone_label = ctk.CTkLabel(self, text="Phone Number:")
        self.phone_label.grid(row=3, column=0, sticky='w', padx=10, pady=10)
        self.phone_entry = ctk.CTkEntry(self)
        self.phone_entry.grid(row=3, column=1, padx=10, pady=10)

        self.email_label = ctk.CTkLabel(self, text="Email:")
        self.email_label.grid(row=4, column=0, sticky='w', padx=10, pady=10)
        self.email_entry = ctk.CTkEntry(self)
        self.email_entry.grid(row=4, column=1, padx=10, pady=10)

        self.occupation_label = ctk.CTkLabel(self, text="Occupation:")
        self.occupation_label.grid(row=5, column=0, sticky='w', padx=10, pady=10)
        self.occupation_entry = ctk.CTkEntry(self)
        self.occupation_entry.grid(row=5, column=1, padx=10, pady=10)

        self.experience_label = ctk.CTkLabel(self, text="HKM Mantra Chanting streak (in years):")
        self.experience_label.grid(row=6, column=0, sticky='w', padx=10, pady=10)
        self.experience_entry = ctk.CTkEntry(self)
        self.experience_entry.grid(row=6, column=1, padx=10, pady=10)

        # Create a button to save the details
        self.save_button = ctk.CTkButton(self, text="Save", command=self.save_details)
        self.save_button.grid(row=7, column=0, columnspan=2, pady=20)

    def get_next_serial_number(self, worksheet):
        """
        Calculate the next serial number (Sl No) by reading the last value in the 'Sl No' column.
        """
        max_row = worksheet.max_row

        # If there is data in the sheet, get the last serial number
        if max_row > 1:
            # Read the last serial number from the 'Sl No' column
            last_sl_no = worksheet.cell(row=max_row, column=1).value
            # Convert the last serial number to an integer
            last_sl_no = int(last_sl_no)
        else:
            last_sl_no = 0

        # Increment the serial number by 1
        next_sl_no = last_sl_no + 1

        return next_sl_no


    def save_details(self):
        # Define the path to the Excel file
        file_path = 'subject_info.xlsx'

        # Load the existing Excel file or create a new workbook if it doesn't exist
        try:
            workbook = load_workbook(file_path)
            worksheet = workbook.active
        except FileNotFoundError:
            # If the file doesn't exist, create a new workbook
            workbook = Workbook()
            worksheet = workbook.active
            # Define headers in the first row
            headers = ["Sl No", "Subject ID", "Name", "Age", "Gender", "PhoneNumber", "Email", "Occupation", "HKM Mantra Chanting streak (in years)"]
            worksheet.append(headers)
            workbook.save(file_path)
            print(f"Created new Excel file: {file_path}")

        # Get the next serial number (Sl No)
        next_sl_no = self.get_next_serial_number(worksheet)

        # Generate a shorter ID using a random number generator
        short_id = f"ID-{random.randint(1000, 9999)}"

        # Collect all user details from the form
        name = self.name_entry.get()
        age = self.age_entry.get()
        gender = self.gender_entry.get()
        phone_number = self.phone_entry.get()
        email = self.email_entry.get()
        occupation = self.occupation_entry.get()
        experience = self.experience_entry.get()

        # Define the new row data including the next serial number and shorter ID
        new_row = [next_sl_no, short_id, name, age, gender, phone_number, email, occupation, experience]

        # Append the new row data to the worksheet
        worksheet.append(new_row)

        # Save the workbook to the file
        workbook.save(file_path)

        print(f"Data saved successfully in '{file_path}'")

        # Provide feedback and animation on successful save
        self.save_button.configure(text="Data Saved!", fg_color="green")
        self.after(2000, self.reset_save_button)

        # Clear the form entries
        self.clear_form()

    def reset_save_button(self):
        # Reset the save button to its original state
        self.save_button.configure(text="Save", fg_color="default_theme_color")

    def clear_form(self):
        # Clear all entry fields
        self.name_entry.delete(0, tk.END)
        self.age_entry.delete(0, tk.END)
        self.gender_entry.delete(0, tk.END)
        self.phone_entry.delete(0, tk.END)
        self.email_entry.delete(0, tk.END)
        self.occupation_entry.delete(0, tk.END)
        self.experience_entry.delete(0, tk.END)

if __name__ == '__main__':
    app = UserForm()
    app.mainloop()
