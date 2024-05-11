from openpyxl import Workbook, load_workbook
import os

folderpath = "files/after"

recorded_file_path = ""

file_count = 0

headers = [
    "Sl No",
    "Subject ID",
    "Name",
    "Age",
    "Gender",
    "PhoneNumber",
    "Email",
    "Occupation",
    "HKM Mantra Chanting streak (in years)",
    "Session start time",
    "Session stop time",
    "Time, sec",
    "IAPF",
    "Baseline Fatigue score",
    "Fatigue score",
    "Baseline Alpha Gravity",
    "Alpha Gravity",
    "Baseline Concentration index",
    "Concentration index",
    "Baseline Relaxation index",
    "Relaxation index",
    "Theta peak frequency",
    "Alpha peak frequency",
    "Beta peak frequency",
    "Chill",
    "Stress",
    "Focus",
    "Anger",
    "Self-control",
]


def run(folder_path):
    """
    Iterates through all .xlsx files in the specified folder and prints their file names.

    Args:
        folder_path (str): Path to the folder containing the .xlsx files.
    """
    # Get the list of all files in the folder
    metric_files = sorted(os.listdir(folder_path))

    print(metric_files)
    # Iterate through the list of files
    for count, file_name in enumerate(metric_files):

        global file_count
        file_count = count

        # Check if the file has a .xlsx extension
        if file_name.endswith('.xlsx'):
            # Print the file name
            file_name = os.path.join(folder_path, file_name)
            print(file_name)
            recorded_file_path = file_name

            global recorded_workbook, sheet

            recorded_workbook = load_workbook(recorded_file_path)
            sheet = recorded_workbook["Sheet1"]

            copy_metrics_into_new_file()

# workbook = load_workbook("generated_excel_file.xlsx")
# worksheet = workbook["Sheet"]


def get_column_by_name(column_name="Beta peak frequency"):

    word_to_search = column_name

    # Variable to hold the data below the specified word
    data_below_word = []

    # Iterate through all rows in the worksheet
    for row in sheet.iter_rows():
        # Iterate through cells in the current row
        for cell in row:
            # Check if the cell contains the word you are looking for
            if cell.value == word_to_search:
                # If the word is found, set a flag to start gathering data
                start_collecting = True
                # Get the row index of the cell where the word is found
                start_row = cell.row + 1

                # Iterate through rows below the found cell's row
                while start_row <= sheet.max_row:
                    # Get the cell in the column of the original word
                    cell_below = sheet.cell(row=start_row, column=cell.column)
                    # Add the cell's value to the list
                    data_below_word.append(cell_below.value)

                    # Move to the next row
                    start_row += 1

                # Break the loop as we have found the word and collected the data
                break

        # Break the outer loop once we find the word
        if len(data_below_word) > 0:
            break

    # Print the data collected below the specified word
    print(f"Data below '{word_to_search}':")
    print(data_below_word)

    return data_below_word


def add_data_under_column(column_name="Beta peak frequency"):
    # Specify the header name to search for
    word_to_search = column_name

    # Your collected data
    data_below_word = get_column_by_name(column_name)

    # Locate the column with the specified header
    header_row = worksheet[1]  # Assuming the header is in the first row
    target_column = None

    # Search for the specified header in the header row
    for cell in header_row:
        if cell.value == word_to_search:
            target_column = cell.column
            break

    # Check if the header was found
    if target_column is not None:
        # Find the last occupied row in the specified column
        last_occupied_row = 1  # Start with the header row
        for row in range(2, worksheet.max_row + 1):
            # Check the cell in the specific column
            cell = worksheet.cell(row=row, column=target_column)
            if cell.value is not None:
                # Update last occupied row if the cell has a value
                last_occupied_row = row
                print("last occupied row:", last_occupied_row)

        # Determine the starting row for adding data (just after the last occupied row)
        start_row = last_occupied_row + 1

        # Add the collected data below the header in the found column
        for index, value in enumerate(data_below_word):
            worksheet.cell(row=start_row + index,
                           column=target_column, value=value)

        # Save the workbook to the file
        save_workbook(workbook)

    else:
        print(
            f"Header '{word_to_search}' not found in the target worksheet '{worksheet.title}'."
        )


def add_headers_to_worksheet():

    # First create all the headings
    worksheet.append(headers)
    return worksheet


def save_workbook(workbook):
    file_path = "generated_excel_file.xlsx"
    workbook.save(file_path)

    print(f"New Excel file saved at: {file_path}")


# add_headers_to_worksheet()
# add_data_under_column("Stress")


def add_session_start_and_stop_time():

    start_time = sheet["B1"].value
    stop_time = sheet["B2"].value

    # Initialize variables to hold the column indices for the specified headers
    sl_no_col = None
    start_time_col = None
    stop_time_col = None

    # Iterate through the header row to find the columns of interest
    header_row = worksheet[1]  # Assuming the header row is the first row

    for cell in header_row:
        if cell.value == "Sl No":
            sl_no_col = cell.column
        elif cell.value == "Session start time":
            start_time_col = cell.column
        elif cell.value == "Session stop time":
            stop_time_col = cell.column

    # Initialize a variable to hold the maximum row number in the "Sl No" column
    max_row_sl_no = 0

    # Iterate through the "Sl No" column to find the maximum occupied row
    for row in range(2, worksheet.max_row + 1):  # Start from row 2 to skip the header
        cell_value = worksheet.cell(row=row, column=sl_no_col).value
        if cell_value is not None:
            max_row_sl_no = row

    # Check if all necessary columns were found
    if sl_no_col and start_time_col and stop_time_col:
        # Add the values of B1 and B2 to the row at max_row_sl_no
        worksheet.cell(row=max_row_sl_no,
                       column=start_time_col, value=start_time)
        worksheet.cell(row=max_row_sl_no,
                       column=stop_time_col, value=stop_time)

        # Save the workbook to the file
        workbook.save("generated_excel_file.xlsx")

        print(
            f"B1 and B2 values have been added to the max row ({max_row_sl_no}) of the columns 'start time' and 'stop time'."
        )

    else:
        print(
            "One or more of the specified columns ('Sl No', 'start time', or 'stop time') were not found in the worksheet."
        )

    # recorded_workbook.close()


# add_session_start_and_stop_time()


def add_serial_num():
    # Initialize a variable to hold the column index of "Sl No"
    sl_no_col = None

    # Locate the "Sl No" column
    header_row = worksheet[1]  # Assuming the header row is the first row
    for cell in header_row:
        if cell.value == "Sl No":
            sl_no_col = cell.column
            break

    # Check if the "Sl No" column was found
    if sl_no_col is None:
        print("The 'Sl No' column was not found in the worksheet.")
        return

    # Calculate the next serial number
    max_row = worksheet.max_row

    # Write the next serial number to the next row in the "Sl No" column
    worksheet.cell(row=max_row + 1, column=sl_no_col, value=file_count)

    # Save the workbook
    workbook.save("generated_excel_file.xlsx")
    print(
        f"Serial number {file_count + 1} has been added to the 'Sl No' column in row {max_row + 1}."
    )


# add_serial_num()


def copy_metrics_into_new_file():
    
    add_serial_num()
    add_session_start_and_stop_time()

    data_headers = [
        "Time, sec",
        "IAPF",
        "Baseline Fatigue score",
        "Fatigue score",
        "Baseline Alpha Gravity",
        "Alpha Gravity",
        "Baseline Concentration index",
        "Concentration index",
        "Baseline Relaxation index",
        "Relaxation index",
        "Theta peak frequency",
        "Alpha peak frequency",
        "Beta peak frequency",
        "Chill",
        "Stress",
        "Focus",
        "Anger",
        "Self-control",
    ]

    for data_header in data_headers:
        add_data_under_column(data_header)
        

# FUNCTIONS COMPLETE HERE


workbook = Workbook()
worksheet = workbook.active

add_headers_to_worksheet()

run(folderpath)

recorded_workbook.close()
workbook.close()

import copy_user_data_into_generated_xlsx