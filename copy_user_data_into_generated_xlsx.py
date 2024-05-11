from openpyxl import load_workbook

input_file = "subject_info.xlsx"
target_file = "generated_excel_file.xlsx"

def copy_matched_rows():
    """
    Reads the data from the original sheet in the Excel file and checks for matching serial numbers
    in the target sheet. If the serial number matches, updates the row in the target sheet.
    """
    # Load the workbooks
    user_workbook = load_workbook(input_file)
    target_workbook = load_workbook(target_file)

    # Access the original sheet and target sheet
    original_sheet = user_workbook.active  # Modify as necessary
    target_sheet = target_workbook.active  # Modify as necessary

    # Create a dictionary to map headers to column indices in the target sheet
    header_to_col_idx = {}
    for col_idx, cell in enumerate(target_sheet[1], start=1):
        header_to_col_idx[cell.value] = col_idx

    # Create a dictionary to map serial numbers to row indices in the target sheet
    target_rows_by_serial_number = {}
    for row_idx, row in enumerate(target_sheet.iter_rows(min_row=2, values_only=True), start=2):
        sl_no = row[0]  # Assuming the first column is "Sl No"
        target_rows_by_serial_number[sl_no] = row_idx

    # Iterate through rows in the original sheet
    for row in original_sheet.iter_rows(min_row=2, values_only=True):
        sl_no = row[0]  # Assuming the first column is "Sl No"

        # Check if the serial number exists in the target sheet
        if sl_no in target_rows_by_serial_number:
            # Get the corresponding row index in the target sheet
            target_row_idx = target_rows_by_serial_number[sl_no]

            # Update the row in the target sheet with the data from the original sheet
            for col_idx, data in enumerate(row, start=1):
                # Find the target column index for the current data point
                target_col_idx = header_to_col_idx[original_sheet.cell(row=1, column=col_idx).value]

                # Update the cell in the target sheet
                target_sheet.cell(row=target_row_idx, column=target_col_idx).value = data

    # Save the workbook
    target_workbook.save(target_file)
    print(f"Matched rows updated in '{target_file}'.")

# Call the function to execute the process
copy_matched_rows()
